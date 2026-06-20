# app/services/staff_service.py
import hashlib
import secrets
from sqlalchemy.orm import Session
from app.database.models import Staff

_MAX_ADMINS = 2


def _hash_password(raw: str) -> str:
    salt = secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac("sha256", raw.encode("utf-8"), salt.encode("utf-8"), 100_000)
    return f"{salt}:{h.hex()}"


def _check_password(raw: str, stored: str) -> bool:
    try:
        salt, hashed = stored.split(":", 1)
        h = hashlib.pbkdf2_hmac("sha256", raw.encode("utf-8"), salt.encode("utf-8"), 100_000)
        return h.hex() == hashed
    except Exception:
        return False


def create_staff(session: Session, name: str,
                 supervisor_id: int | None = None,
                 is_department_head: bool = False,
                 email: str = "") -> Staff:
    staff = Staff(name=name, supervisor_id=supervisor_id,
                  is_department_head=is_department_head,
                  email=email or None)
    session.add(staff)
    session.commit()
    session.refresh(staff)
    return staff


def get_staff(session: Session, staff_id: int) -> Staff | None:
    return session.get(Staff, staff_id)


def get_active_staff(session: Session) -> list[Staff]:
    return session.query(Staff).filter_by(is_active=True).order_by(Staff.name).all()


def get_all_staff(session: Session) -> list[Staff]:
    return session.query(Staff).order_by(Staff.name).all()


def deactivate_staff(session: Session, staff_id: int) -> None:
    staff = session.get(Staff, staff_id)
    if staff:
        staff.is_active = False
        session.commit()


def reactivate_staff(session: Session, staff_id: int) -> None:
    staff = session.get(Staff, staff_id)
    if staff:
        staff.is_active = True
        session.commit()


def update_staff_name(session: Session, staff_id: int, name: str) -> Staff:
    staff = session.get(Staff, staff_id)
    staff.name = name
    session.commit()
    return staff


def update_staff(session: Session, staff_id: int, name: str,
                 supervisor_id: int | None = None,
                 is_department_head: bool = False,
                 email: str = "") -> Staff:
    staff = session.get(Staff, staff_id)
    staff.name = name
    staff.supervisor_id = supervisor_id
    staff.is_department_head = is_department_head
    staff.email = email.strip() or None
    session.commit()
    return staff


def get_department_heads(session: Session) -> list[Staff]:
    return (session.query(Staff)
            .filter_by(is_department_head=True, is_active=True)
            .order_by(Staff.name)
            .all())


def import_staff_from_csv(session: Session, file_path: str) -> tuple[int, int]:
    """CSVから職員を追加インポートする。同名の職員はスキップ。

    対応ヘッダー（大文字小文字・スペース不問）:
      氏名 / 名前 / name
      メール / メールアドレス / email
      所属長 / 所属長フラグ / is_head  →  ○/1/true で所属長フラグON
      担当所属長 / 上司 / supervisor     →  既存職員の氏名で参照

    Returns: (追加件数, スキップ件数)
    """
    import csv, pathlib

    suffix = pathlib.Path(file_path).suffix.lower()
    rows: list[dict] = []
    if suffix in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
        for excel_row in ws.iter_rows(min_row=2, values_only=True):
            rows.append({h: (str(v).strip() if v is not None else "") for h, v in zip(headers, excel_row)})
        wb.close()
    else:
        with open(file_path, encoding="utf-8-sig", newline="") as f:
            rows = list(csv.DictReader(f))

    _NAME_KEYS  = {"氏名", "名前", "name"}
    _EMAIL_KEYS = {"メール", "メールアドレス", "email"}
    _HEAD_KEYS  = {"所属長", "所属長フラグ", "is_head"}
    _SUP_KEYS   = {"担当所属長", "上司", "supervisor"}

    def _find(row: dict, keys: set) -> str:
        for k in row:
            if k.lower().replace(" ", "") in {x.lower().replace(" ", "") for x in keys}:
                return row[k].strip()
        return ""

    existing = {s.name: s for s in session.query(Staff).all()}
    added = skipped = 0

    for row in rows:
        name = _find(row, _NAME_KEYS)
        if not name:
            continue
        if name in existing:
            skipped += 1
            continue

        email      = _find(row, _EMAIL_KEYS)
        head_raw   = _find(row, _HEAD_KEYS)
        is_head    = head_raw.lower() in ("○", "1", "true", "yes", "はい")
        sup_name   = _find(row, _SUP_KEYS)
        sup_id     = existing[sup_name].id if sup_name and sup_name in existing else None

        staff = Staff(name=name, email=email or None,
                      is_department_head=is_head, supervisor_id=sup_id)
        session.add(staff)
        session.flush()
        existing[name] = staff
        added += 1

    session.commit()
    return added, skipped


# ── パスワード管理 ────────────────────────────────────────

def set_password(session: Session, staff_id: int, raw_password: str) -> None:
    """パスワードをハッシュ化して保存する。"""
    staff = session.get(Staff, staff_id)
    if staff:
        staff.password_hash = _hash_password(raw_password)
        session.commit()


def verify_password(session: Session, staff_id: int, raw_password: str) -> bool:
    """パスワードを照合する。"""
    staff = session.get(Staff, staff_id)
    if not staff or not staff.password_hash:
        return False
    return _check_password(raw_password, staff.password_hash)


def reset_password(session: Session, staff_id: int) -> None:
    """パスワードをリセット（NULL）する。次回ログイン時に再設定を求める。"""
    staff = session.get(Staff, staff_id)
    if staff:
        staff.password_hash = None
        session.commit()


# ── 管理者管理 ────────────────────────────────────────────

def count_admins(session: Session) -> int:
    return session.query(Staff).filter_by(is_admin=True, is_active=True).count()


def has_any_admin(session: Session) -> bool:
    return count_admins(session) > 0


def set_admin(session: Session, staff_id: int, is_admin: bool) -> None:
    """管理者フラグを設定する。管理者は最大2名まで。"""
    if is_admin and count_admins(session) >= _MAX_ADMINS:
        raise ValueError(f"管理者は最大{_MAX_ADMINS}名まで設定できます。")
    staff = session.get(Staff, staff_id)
    if staff:
        staff.is_admin = is_admin
        session.commit()


def set_department_head(session: Session, staff_id: int,
                        is_department_head: bool) -> None:
    staff = session.get(Staff, staff_id)
    if staff:
        staff.is_department_head = is_department_head
        session.commit()


def update_staff_email(session: Session, staff_id: int, email: str) -> None:
    staff = session.get(Staff, staff_id)
    if staff:
        staff.email = email.strip() or None
        session.commit()
