# app/services/report_service.py
import os
from sqlalchemy.orm import Session
from app.database.models import Issuance, ProjectMember, Project, Payment
from app.services.project_service import get_project_progress


def get_unpaid_report(session: Session,
                      fiscal_year: int | None = None,
                      project_id: int | None = None) -> list[dict]:
    q = (session.query(Issuance, Project)
         .join(Project, Issuance.project_id == Project.id)
         .filter(Issuance.status.in_(["準備中", "発行済み"])))
    if fiscal_year:
        q = q.filter(Project.fiscal_year == fiscal_year)
    if project_id:
        q = q.filter(Issuance.project_id == project_id)

    rows = []
    for iss, proj in q.all():
        pm = session.get(ProjectMember, iss.project_member_id) if iss.project_member_id else None
        description = "、".join(l.item_name for l in iss.lines if l.item_name)
        rows.append({
            "doc_number":          iss.doc_number,
            "project_name":        proj.name,
            "fiscal_year":         proj.fiscal_year,
            "organization_name":   iss.recipient_organization or (pm.organization_name if pm else ""),
            "representative_name": iss.recipient_name or (pm.representative_name if pm else ""),
            "description":         description,
            "member_number":       "",
            "amount":              int(iss.amount),
            "status":              iss.status,
            "doc_type":            iss.doc_type,
        })
    return rows


def get_payment_report(session: Session,
                       fiscal_year: int | None = None,
                       project_id: int | None = None) -> list[dict]:
    q = (session.query(Payment, Issuance, Project)
         .join(Issuance, Payment.issuance_id == Issuance.id)
         .join(Project, Issuance.project_id == Project.id))
    if fiscal_year:
        q = q.filter(Project.fiscal_year == fiscal_year)
    if project_id:
        q = q.filter(Issuance.project_id == project_id)
    q = q.order_by(Payment.payment_date.desc())

    rows = []
    for payment, iss, proj in q.all():
        description = "、".join(l.item_name for l in iss.lines if l.item_name)
        rows.append({
            "payment_date":   payment.payment_date.strftime("%Y/%m/%d"),
            "doc_number":     iss.doc_number,
            "project_name":   proj.name,
            "fiscal_year":    proj.fiscal_year,
            "organization":   iss.recipient_organization or iss.recipient_name,
            "description":    description,
            "amount":         int(payment.amount),
            "payment_method": payment.payment_method,
            "staff_name":     payment.staff_name,
        })
    return rows


def get_project_summary(session: Session,
                         fiscal_year: int | None = None) -> list[dict]:
    q = session.query(Project).filter(Project.status.in_(["active", "closed"]))
    if fiscal_year:
        q = q.filter(Project.fiscal_year == fiscal_year)

    rows = []
    for proj in q.order_by(Project.fiscal_year.desc(), Project.name).all():
        p = get_project_progress(session, proj.id)
        total_amount = sum(int(iss.amount) for iss in
                           session.query(Issuance).filter_by(project_id=proj.id).all())
        paid_amount  = sum(int(iss.amount) for iss in
                           session.query(Issuance).filter_by(
                               project_id=proj.id, status="支払済み").all())
        rows.append({
            "fiscal_year":    proj.fiscal_year,
            "project_name":   proj.name,
            "project_type":   "名簿あり" if proj.project_type == "list" else "その場入力",
            "total":          p["total"],
            "invoice_issued": p["invoice_issued"],
            "receipt_issued": p["receipt_issued"],
            "pending":        p["pending"],
            "total_amount":   total_amount,
            "paid_amount":    paid_amount,
        })
    return rows


def export_to_excel(rows: list[dict], headers: list[str],
                    output_path: str) -> str:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E40AF")
        cell.alignment = Alignment(horizontal="center")
    keys = list(rows[0].keys()) if rows else []
    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return output_path
