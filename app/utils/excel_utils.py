# app/utils/excel_utils.py
import openpyxl

MEMBER_COLUMNS = [
    "member_number",
    "organization_name",
    "organization_kana",
    "representative_name",
    "representative_kana",
    "department",
    "postal_code",
    "address",
    "address2",
    "phone",
    "email",
]

ROSTER_COLUMNS = MEMBER_COLUMNS

# 取り込み先フィールド → 日本語ラベル（MEMBER_COLUMNS と同順）
FIELD_LABELS = {
    "member_number": "会員番号",
    "organization_name": "事業所名",
    "organization_kana": "フリガナ",
    "representative_name": "代表者名",
    "representative_kana": "代表者フリガナ",
    "department": "所属・役職名",
    "postal_code": "郵便番号",
    "address": "住所１",
    "address2": "住所２",
    "phone": "電話",
    "email": "メール",
}

# 必須（このいずれかが空の行は取り込まない）
REQUIRED_ANY = ("organization_name", "representative_name")


def parse_tsv_text(text: str) -> list[dict]:
    """ExcelからコピーしたTSVテキストを会員データのリストに変換する"""
    rows = []
    for line in text.splitlines():
        if not line.strip():
            continue
        cells = line.split("\t")
        while len(cells) < len(MEMBER_COLUMNS):
            cells.append("")
        row = {col: cells[i].strip() for i, col in enumerate(MEMBER_COLUMNS)}
        if not row["organization_name"] and not row["representative_name"]:
            continue
        rows.append(row)
    return rows


def parse_excel_file(file_path: str, sheet_name: str | None = None,
                     header_row: int = 1) -> list[dict]:
    """Excelファイルを読み込んで会員データのリストに変換する"""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < header_row:
            continue
        cells = [str(c).strip() if c is not None else "" for c in row]
        while len(cells) < len(MEMBER_COLUMNS):
            cells.append("")
        data = {col: cells[j] for j, col in enumerate(MEMBER_COLUMNS)}
        if not data["organization_name"] and not data["representative_name"]:
            continue
        rows.append(data)
    wb.close()
    return rows


# ── 列マッピング方式（位置固定ではなく、列を選んで対応づける）──────────────

def parse_tsv_text_raw(text: str) -> list[list[str]]:
    """貼り付けたTSVを「行ごとのセル配列」に変換（フィールド割り当てはしない）。"""
    rows = []
    for line in text.splitlines():
        if not line.strip():
            continue
        rows.append([c.strip() for c in line.split("\t")])
    return rows


def parse_excel_file_raw(file_path: str, sheet_name: str | None = None) -> list[list[str]]:
    """Excelを「行ごとのセル配列」に変換（フィールド割り当てはしない）。"""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if not any(cells):
            continue
        rows.append(cells)
    wb.close()
    return rows


def column_count(raw_rows: list[list[str]]) -> int:
    """セル配列の最大列数。"""
    return max((len(r) for r in raw_rows), default=0)


def default_positional_mapping(num_cols: int) -> dict[str, int | None]:
    """従来どおり左から順に割り当てた初期マッピング。"""
    return {field: (i if i < num_cols else None)
            for i, field in enumerate(MEMBER_COLUMNS)}


def guess_mapping_from_header(header_cells: list[str]) -> dict[str, int | None]:
    """見出し行の文字列からフィールドを推測して割り当てる。

    日本語ラベル（事業所名 等）またはフィールド名（organization_name 等）と
    一致した列を対応づける。一致しないフィールドは None。
    """
    label_to_field = {label: field for field, label in FIELD_LABELS.items()}
    mapping: dict[str, int | None] = {field: None for field in MEMBER_COLUMNS}
    for i, raw in enumerate(header_cells):
        h = (raw or "").strip()
        if h in label_to_field:
            mapping[label_to_field[h]] = i
        elif h in MEMBER_COLUMNS:
            mapping[h] = i
    return mapping


def build_member_rows(raw_rows: list[list[str]],
                      mapping: dict[str, int | None],
                      has_header: bool = False) -> list[dict]:
    """マッピングに従って会員データ（dict）のリストを組み立てる。

    has_header=True なら先頭行を見出しとして除外。
    必須（事業所名・代表者名のいずれか）が空の行は除外。
    """
    data_rows = raw_rows[1:] if has_header else raw_rows
    result = []
    for cells in data_rows:
        row = {}
        for field in MEMBER_COLUMNS:
            idx = mapping.get(field)
            if idx is None or idx >= len(cells):
                row[field] = ""
            else:
                row[field] = cells[idx].strip()
        if not any(row[f] for f in REQUIRED_ANY):
            continue
        result.append(row)
    return result
