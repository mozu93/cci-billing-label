# app/ui/issuance_from_project.py
import calendar
import os
from datetime import date
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QPushButton, QLabel, QHeaderView, QComboBox, QLineEdit, QMessageBox,
    QSpinBox, QFileDialog, QProgressDialog, QCheckBox, QDateEdit
)
from PyQt6.QtCore import Qt, QTimer, QDate
from app.database.connection import get_session
from app.services.project_service import (
    get_projects, get_project_members, get_project_templates
)
from app.services.category_service import get_active_categories
from app.services.issuance_service import create_issuance_for_member, mark_as_issued
from app.utils import current_user


COL_CHK  = 0
COL_NUM  = 1   # 会員番号
COL_ORG  = 2   # 事業所名
COL_KANA = 3   # フリガナ
COL_REP  = 4   # 代表者名
COL_PROJ = 5   # 件名（すべて選択時専用）
# 数量列: 5 〜 5+len(templates)-1  ※件名選択時
# 請求書列 = 5+len(templates)、領収書列 = 6+len(templates) — テンプレート数で可変


class _QtySpinBox(QSpinBox):
    """テーブル内数量入力用: Enter で同列の次行へフォーカスを移動する。"""

    def __init__(self, table: "QTableWidget", row: int, col: int):
        super().__init__()
        self._tbl = table
        self._row = row
        self._col = col

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key.Key_Return, Qt.Key.Key_Enter):
            next_row = self._row + 1
            if next_row < self._tbl.rowCount():
                nxt = self._tbl.cellWidget(next_row, self._col)
                if nxt:
                    nxt.setFocus()
                    nxt.selectAll()
        else:
            super().keyPressEvent(event)


class _CheckableTable(QTableWidget):
    """チェックボックス列の Shift+クリック範囲選択に対応したテーブル。"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._last_checked_row: int = -1

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            idx = self.indexAt(event.pos())
            if idx.isValid() and idx.column() == COL_CHK:
                item = self.item(idx.row(), COL_CHK)
                if item and (item.flags() & Qt.ItemFlag.ItemIsUserCheckable):
                    if (event.modifiers() & Qt.KeyboardModifier.ShiftModifier
                            and self._last_checked_row >= 0):
                        new_state = (Qt.CheckState.Unchecked
                                     if item.checkState() == Qt.CheckState.Checked
                                     else Qt.CheckState.Checked)
                        r1 = min(self._last_checked_row, idx.row())
                        r2 = max(self._last_checked_row, idx.row())
                        self.blockSignals(True)
                        for r in range(r1, r2 + 1):
                            it = self.item(r, COL_CHK)
                            if it:
                                it.setCheckState(new_state)
                        self.blockSignals(False)
                        self._last_checked_row = idx.row()
                        return
                    else:
                        self._last_checked_row = idx.row()
        super().mousePressEvent(event)


class IssuanceFromProjectWidget(QWidget):
    def __init__(self, doc_type: str = "invoice"):
        super().__init__()
        self._doc_type = doc_type
        self._templates: list[dict] = []  # [{id, name}, ...]
        self._sort_col: int = -1   # -1 = 登録順（sort_order）
        self._sort_asc: bool = True
        self._qty_cache: dict[int, dict[int, int]] = {}    # {pm_id: {tmpl_id: qty}}
        self._price_cache: dict[int, dict[int, int]] = {}  # {pm_id: {tmpl_id: unit_price}}
        self._all_projects: list = []
        self._build()
        self._load_projects()

    @property
    def _is_all_mode(self) -> bool:
        return self._proj_combo.count() > 0 and self._proj_combo.currentData() is None

    @property
    def _col_inv(self) -> int:
        return 6 if self._is_all_mode else 5 + len(self._templates) * 2

    @property
    def _col_rcp(self) -> int:
        return 7 if self._is_all_mode else 6 + len(self._templates) * 2

    def _build(self):
        layout = QVBoxLayout(self)

        # ── フィルタ行（年度 / 業務区分 / 件名 / 表示 / 検索） ────────
        filter_row = QHBoxLayout()
        filter_row.addWidget(QLabel("年度："))
        self._year_combo = QComboBox()
        self._year_combo.setMinimumWidth(95)
        self._year_combo.currentIndexChanged.connect(self._filter_projects)
        filter_row.addWidget(self._year_combo)
        filter_row.addWidget(QLabel("業務区分："))
        self._cat_combo = QComboBox()
        self._cat_combo.setMinimumWidth(120)
        self._cat_combo.currentIndexChanged.connect(self._filter_projects)
        filter_row.addWidget(self._cat_combo)
        filter_row.addWidget(QLabel("件名："))
        self._proj_combo = QComboBox()
        self._proj_combo.setMinimumWidth(200)
        self._proj_combo.currentIndexChanged.connect(self._on_project_changed)
        filter_row.addWidget(self._proj_combo)
        filter_row.addWidget(QLabel("表示："))
        self._filter_combo = QComboBox()
        self._filter_combo.addItems(["未発行のみ", "すべて"])
        self._filter_combo.currentIndexChanged.connect(self._load_members)
        filter_row.addWidget(self._filter_combo)
        filter_row.addSpacing(12)
        self._search = QLineEdit()
        self._search.setPlaceholderText("事業所名・代表者名で絞り込み")
        self._timer = QTimer()
        self._timer.setSingleShot(True)
        self._timer.timeout.connect(self._load_members)
        self._search.textChanged.connect(lambda: self._timer.start(300))
        filter_row.addWidget(self._search)
        layout.addLayout(filter_row)

        # ── アクション行（発行 / 配付方法 / 支払期日 / 封筒 / Excel） ─
        label = "請求書" if self._doc_type == "invoice" else "領収書"
        action_row = QHBoxLayout()
        self._btn_issue = QPushButton(f"選択した{label}を発行")
        self._btn_issue.clicked.connect(self._issue_checked)
        self._delivery_combo = QComboBox()
        self._delivery_combo.addItems(["窓口手渡し", "郵送", "メール送付"])
        action_row.addWidget(self._btn_issue)
        action_row.addWidget(QLabel("配付方法："))
        action_row.addWidget(self._delivery_combo)

        if self._doc_type == "invoice":
            today = date.today()
            nm_year = today.year + 1 if today.month == 12 else today.year
            nm_month = 1 if today.month == 12 else today.month + 1
            last_day = calendar.monthrange(nm_year, nm_month)[1]
            self._due_date = QDateEdit(QDate(nm_year, nm_month, last_day))
            self._due_date.setCalendarPopup(True)
            self._due_date.setDisplayFormat("yyyy/MM/dd")
            self._window_envelope_chk = QCheckBox("窓あき封筒モード")
            self._show_person_chk = QCheckBox("役職名・氏名を印字")
            self._show_person_chk.setChecked(True)
            action_row.addSpacing(16)
            action_row.addWidget(QLabel("支払期日："))
            action_row.addWidget(self._due_date)
            action_row.addSpacing(8)
            action_row.addWidget(self._window_envelope_chk)
            action_row.addSpacing(4)
            action_row.addWidget(self._show_person_chk)
        else:
            today = date.today()
            self._issued_date = QDateEdit(QDate(today.year, today.month, today.day))
            self._issued_date.setCalendarPopup(True)
            self._issued_date.setDisplayFormat("yyyy/MM/dd")
            action_row.addSpacing(16)
            action_row.addWidget(QLabel("発行日："))
            action_row.addWidget(self._issued_date)

        action_row.addStretch()
        self._btn_export_xlsx = QPushButton("Excel出力")
        self._btn_export_xlsx.setToolTip(
            "表示中の名簿と数量をExcelに出力します。\n"
            "Excelで数量を入力後、「Excel取込」で読み込めます。")
        self._btn_export_xlsx.clicked.connect(self._export_excel)
        action_row.addWidget(self._btn_export_xlsx)
        self._btn_import_xlsx = QPushButton("Excel取込")
        self._btn_import_xlsx.setToolTip(
            "Excel出力したファイルを読み込み、数量と発行チェックを画面に反映します。")
        self._btn_import_xlsx.clicked.connect(self._import_excel)
        action_row.addWidget(self._btn_import_xlsx)
        layout.addLayout(action_row)

        self._table = _CheckableTable(0, 7)
        self._table.horizontalHeader().sectionClicked.connect(self._on_header_clicked)
        self._table.horizontalHeader().setSortIndicatorShown(True)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.cellDoubleClicked.connect(self._on_row_double_clicked)
        # ヘッダー左端に本物のチェックボックスを配置
        self._header_chk = QCheckBox(self._table.horizontalHeader())
        self._header_chk.setTristate(False)
        self._header_chk.toggled.connect(self._on_header_checkbox_toggled)
        self._table.horizontalHeader().sectionResized.connect(
            lambda _l, _o, _n: self._reposition_header_chk())
        self._table.setHorizontalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self._table.setVerticalScrollBarPolicy(
            Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self._setup_table_columns()
        layout.addWidget(self._table)

        self._status_label = QLabel("")
        layout.addWidget(self._status_label)

    def _setup_table_columns(self):
        hdr = self._table.horizontalHeader()
        fixed       = QHeaderView.ResizeMode.Fixed
        interactive = QHeaderView.ResizeMode.Interactive
        rtc         = QHeaderView.ResizeMode.ResizeToContents

        if self._is_all_mode:
            self._table.setColumnCount(8)
            self._table.setHorizontalHeaderLabels(
                ["", "会員番号", "事業所名", "フリガナ", "代表者名", "件名", "請求書", "領収書"])
            hdr.setSectionResizeMode(COL_CHK,  fixed);      self._table.setColumnWidth(COL_CHK,  30)
            hdr.setSectionResizeMode(COL_NUM,  interactive); self._table.setColumnWidth(COL_NUM,  80)
            hdr.setSectionResizeMode(COL_ORG,  interactive); self._table.setColumnWidth(COL_ORG, 180)
            hdr.setSectionResizeMode(COL_KANA, interactive); self._table.setColumnWidth(COL_KANA,140)
            hdr.setSectionResizeMode(COL_REP,  interactive); self._table.setColumnWidth(COL_REP, 100)
            hdr.setSectionResizeMode(COL_PROJ, interactive); self._table.setColumnWidth(COL_PROJ,200)
            hdr.setSectionResizeMode(6, rtc)
            hdr.setSectionResizeMode(7, rtc)
        else:
            n = len(self._templates)
            self._table.setColumnCount(7 + n * 2)
            headers = ["", "会員番号", "事業所名", "フリガナ", "代表者名"]
            for tmpl in self._templates:
                headers.append(f"{tmpl['name']}\n単価")
                headers.append(f"{tmpl['name']}\n数量")
            headers += ["請求書", "領収書"]
            self._table.setHorizontalHeaderLabels(headers)
            hdr.setSectionResizeMode(COL_CHK,  fixed);      self._table.setColumnWidth(COL_CHK,  30)
            hdr.setSectionResizeMode(COL_NUM,  interactive); self._table.setColumnWidth(COL_NUM,  80)
            hdr.setSectionResizeMode(COL_ORG,  interactive); self._table.setColumnWidth(COL_ORG, 180)
            hdr.setSectionResizeMode(COL_KANA, interactive); self._table.setColumnWidth(COL_KANA,140)
            hdr.setSectionResizeMode(COL_REP,  interactive); self._table.setColumnWidth(COL_REP, 100)
            for i in range(n):
                hdr.setSectionResizeMode(5 + i * 2, interactive)
                self._table.setColumnWidth(5 + i * 2, 80)
                hdr.setSectionResizeMode(5 + i * 2 + 1, interactive)
                self._table.setColumnWidth(5 + i * 2 + 1, 100)
            for col in (self._col_inv, self._col_rcp):
                hdr.setSectionResizeMode(col, rtc)

        if self._sort_col >= 0:
            hdr.setSortIndicator(
                self._sort_col,
                Qt.SortOrder.AscendingOrder if self._sort_asc else Qt.SortOrder.DescendingOrder)
        else:
            hdr.setSortIndicator(-1, Qt.SortOrder.AscendingOrder)
        self._reposition_header_chk()

    # ── ヘッダークリック：全選択 / 全解除 ─────────────────────────

    def _on_header_clicked(self, col: int):
        if col == COL_CHK:
            return  # QCheckBox ウィジェットが処理
        # 数量・単価 SpinBox 列はソート対象外
        spin_cols = set(range(5, 5 + len(self._templates) * 2))
        if col in spin_cols:
            return
        self._save_qty_cache()
        if self._sort_col == col:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = col
            self._sort_asc = True
        self._load_members()

    def _reposition_header_chk(self):
        hdr = self._table.horizontalHeader()
        x = hdr.sectionViewportPosition(COL_CHK)
        w = hdr.sectionSize(COL_CHK)
        h = hdr.height()
        chk = self._header_chk
        chk.resize(chk.sizeHint())
        chk.move(x + (w - chk.width()) // 2, (h - chk.height()) // 2)
        chk.show()

    def _on_header_checkbox_toggled(self, checked: bool):
        state = Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked
        self._table.blockSignals(True)
        for r in range(self._table.rowCount()):
            it = self._table.item(r, COL_CHK)
            if it:
                it.setCheckState(state)
        self._table.blockSignals(False)

    def _on_row_double_clicked(self, row: int, col: int):
        if not self._is_all_mode:
            return
        proj_item = self._table.item(row, COL_PROJ)
        if not proj_item:
            return
        proj_name = proj_item.text()
        for i in range(self._proj_combo.count()):
            if self._proj_combo.itemText(i) == proj_name:
                self._proj_combo.setCurrentIndex(i)
                return

    # ── プロジェクト読み込み ───────────────────────────────────────

    def showEvent(self, event):
        super().showEvent(event)
        self._load_projects()
        self._reposition_header_chk()

    def _load_projects(self):
        session = get_session()
        try:
            self._all_projects = get_projects(session, status="active")
            cats = get_active_categories(session)
        finally:
            session.close()

        # 年度コンボ（重複なし降順）
        years = sorted({p.fiscal_year for p in self._all_projects}, reverse=True)
        current_year = self._year_combo.currentData()
        self._year_combo.blockSignals(True)
        self._year_combo.clear()
        self._year_combo.addItem("すべて", None)
        for y in years:
            self._year_combo.addItem(f"{y}年度", y)
        # デフォルト：最新年度
        if years and current_year is None:
            self._year_combo.setCurrentIndex(1)
        else:
            for i in range(self._year_combo.count()):
                if self._year_combo.itemData(i) == current_year:
                    self._year_combo.setCurrentIndex(i)
                    break
        self._year_combo.blockSignals(False)

        # 業務区分コンボ：_all_projects に含まれるカテゴリのみ表示（窓口除外済み）
        used_cat_ids = {p.category_id for p in self._all_projects}
        current_cat = self._cat_combo.currentData()
        self._cat_combo.blockSignals(True)
        self._cat_combo.clear()
        self._cat_combo.addItem("すべて", None)
        for c in cats:
            if c.id in used_cat_ids:
                self._cat_combo.addItem(c.name, c.id)
        for i in range(self._cat_combo.count()):
            if self._cat_combo.itemData(i) == current_cat:
                self._cat_combo.setCurrentIndex(i)
                break
        self._cat_combo.blockSignals(False)

        self._filter_projects()

    def _filter_projects(self):
        sel_year = self._year_combo.currentData()
        sel_cat  = self._cat_combo.currentData()
        current_id = self._proj_combo.currentData()
        is_first_load = self._proj_combo.count() == 0
        self._proj_combo.blockSignals(True)
        self._proj_combo.clear()
        self._proj_combo.addItem("すべて", None)
        for p in self._all_projects:
            if sel_year is not None and p.fiscal_year != sel_year:
                continue
            if sel_cat is not None and p.category_id != sel_cat:
                continue
            self._proj_combo.addItem(p.name, p.id)
        if is_first_load:
            # 初回：最初のプロジェクトを自動選択（あれば）
            if self._proj_combo.count() > 1:
                self._proj_combo.setCurrentIndex(1)
        else:
            # 以前の選択を復元（current_id が None なら「すべて」のまま）
            for i in range(self._proj_combo.count()):
                if self._proj_combo.itemData(i) == current_id:
                    self._proj_combo.setCurrentIndex(i)
                    break
        self._proj_combo.blockSignals(False)
        self._on_project_changed()

    def _on_project_changed(self):
        project_id = self._proj_combo.currentData()
        is_all = project_id is None
        if is_all:
            self._templates = []
        else:
            session = get_session()
            try:
                pts = get_project_templates(session, project_id)
                self._templates = [
                    {
                        "id": pt.item_template.id,
                        "name": pt.item_template.name,
                        "unit_price": int(pt.unit_price_override or pt.item_template.unit_price or 0),
                        "default_qty": int(pt.default_quantity or 1),
                    }
                    for pt in pts
                ]
            finally:
                session.close()
        for btn in (self._btn_issue, self._btn_export_xlsx, self._btn_import_xlsx):
            btn.setEnabled(not is_all)
        self._btn_issue.setToolTip("件名を選択すると発行できます" if is_all else "")
        self._setup_table_columns()
        self._load_members()

    # ── メンバー一覧読み込み ──────────────────────────────────────

    _STATUS_SHORT = {"発行済み": "発行済", "支払済み": "支払済", "準備中": "準備中"}

    def _cell_text(self, iss) -> str:
        if iss is None:
            return "未発行"
        short = self._STATUS_SHORT.get(iss.status, iss.status)
        return f"{short} {iss.doc_number}".strip()

    def _load_members(self):
        project_id = self._proj_combo.currentData()
        query    = self._search.text().strip().lower()
        show_all = self._filter_combo.currentIndex() == 1
        doc_type = self._doc_type
        is_all   = self._is_all_mode

        if is_all:
            project_ids   = [self._proj_combo.itemData(i)
                             for i in range(self._proj_combo.count())
                             if self._proj_combo.itemData(i) is not None]
            proj_name_map = {self._proj_combo.itemData(i): self._proj_combo.itemText(i)
                             for i in range(1, self._proj_combo.count())}
        else:
            project_ids   = [project_id]
            proj_name_map = {}

        session = get_session()
        try:
            from app.database.models import Issuance
            pm_data = []
            issued_count = 0
            for pid in project_ids:
                for pm in get_project_members(session, pid):
                    inv = (session.query(Issuance)
                           .filter_by(project_member_id=pm.id, doc_type="invoice")
                           .order_by(Issuance.created_at.desc())
                           .first())
                    rcp = (session.query(Issuance)
                           .filter_by(project_member_id=pm.id, doc_type="receipt")
                           .order_by(Issuance.created_at.desc())
                           .first())
                    voided = inv is None and rcp is not None
                    sel = inv if doc_type == "invoice" else rcp
                    sel_status = sel.status if sel else "未発行"
                    hide_issued = sel_status in ("発行済み", "支払済み")
                    hide_voided = doc_type == "invoice" and voided
                    if hide_issued:
                        issued_count += 1
                    if not show_all and (hide_issued or hide_voided):
                        continue
                    if query:
                        targets = [
                            pm.organization_name or "",
                            pm.representative_name or "",
                            pm.organization_kana or "",
                        ]
                        if not any(query in t.lower() for t in targets):
                            continue
                    inv_text = "無効" if voided else self._cell_text(inv)
                    pm_data.append((
                        pm.id, pm,
                        inv_text, self._cell_text(rcp),
                        inv.id if inv else None, rcp.id if rcp else None,
                        proj_name_map.get(pid, ""),
                    ))
        finally:
            session.close()

        # ソート
        col_inv = self._col_inv
        col_rcp = self._col_rcp
        sc = self._sort_col

        def _key(item):
            _, pm, inv_text, rcp_text, _, _, proj_name = item
            if sc < 0:                         return pm.sort_order or 0
            if sc == COL_NUM:                  return pm.member_number or ""
            if sc == COL_ORG:                  return pm.organization_name or ""
            if sc == COL_KANA:                 return pm.organization_kana or ""
            if sc == COL_REP:                  return pm.representative_name or ""
            if is_all and sc == COL_PROJ:      return proj_name
            if sc == col_inv:                  return inv_text
            if sc == col_rcp:                  return rcp_text
            return ""

        pm_data.sort(key=_key, reverse=(sc >= 0 and not self._sort_asc))

        self._table._last_checked_row = -1
        self._header_chk.blockSignals(True)
        self._header_chk.setChecked(False)
        self._header_chk.blockSignals(False)

        self._table.setRowCount(0)
        for pm_id, pm, inv_text, rcp_text, inv_id, rcp_id, proj_name in pm_data:
            row = self._table.rowCount()
            self._table.insertRow(row)

            chk_item = QTableWidgetItem()
            chk_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsUserCheckable)
            chk_item.setCheckState(Qt.CheckState.Unchecked)
            self._table.setItem(row, COL_CHK, chk_item)

            row_data = (pm_id, inv_id, rcp_id)
            fixed_cols = [
                (COL_NUM,  pm.member_number or ""),
                (COL_ORG,  pm.organization_name or ""),
                (COL_KANA, pm.organization_kana or ""),
                (COL_REP,  pm.representative_name or ""),
            ]
            if is_all:
                fixed_cols.append((COL_PROJ, proj_name))
            fixed_cols += [(col_inv, inv_text), (col_rcp, rcp_text)]
            for col, val in fixed_cols:
                it = QTableWidgetItem(val)
                it.setData(Qt.ItemDataRole.UserRole, row_data)
                self._table.setItem(row, col, it)

            for col_offset, tmpl in enumerate(self._templates):
                price_col = 5 + col_offset * 2
                qty_col   = 5 + col_offset * 2 + 1
                _base = "QSpinBox { min-height: 0; padding: 1px 4px; }"
                _mod  = "QSpinBox { min-height: 0; padding: 1px 4px; background: #FFF9C4; }"

                default_price = tmpl["unit_price"]
                price_spin = _QtySpinBox(self._table, row, price_col)
                price_spin.setRange(0, 9999999)
                cached_price = self._price_cache.get(pm_id, {}).get(tmpl["id"], default_price)
                price_spin.setValue(cached_price)
                price_spin.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                price_spin.setStyleSheet(_mod if cached_price != default_price else _base)

                def _on_price(v, pid=pm_id, tid=tmpl["id"], dp=default_price,
                              sp=price_spin, b=_base, m=_mod):
                    self._price_cache.setdefault(pid, {})[tid] = v
                    sp.setStyleSheet(m if v != dp else b)

                price_spin.valueChanged.connect(_on_price)
                self._table.setCellWidget(row, price_col, price_spin)

                default_qty = tmpl["default_qty"]
                spin = _QtySpinBox(self._table, row, qty_col)
                spin.setRange(0, 9999)
                cached_qty = self._qty_cache.get(pm_id, {}).get(tmpl["id"], default_qty)
                spin.setValue(cached_qty)
                spin.setAlignment(Qt.AlignmentFlag.AlignCenter)
                spin.setStyleSheet(_mod if cached_qty != default_qty else _base)

                def _on_qty(v, pid=pm_id, tid=tmpl["id"], dq=default_qty,
                            sp=spin, b=_base, m=_mod):
                    self._qty_cache.setdefault(pid, {})[tid] = v
                    sp.setStyleSheet(m if v != dq else b)

                spin.valueChanged.connect(_on_qty)
                self._table.setCellWidget(row, qty_col, spin)

        hdr = self._table.horizontalHeader()
        if self._sort_col >= 0:
            hdr.setSortIndicator(
                self._sort_col,
                Qt.SortOrder.AscendingOrder if self._sort_asc else Qt.SortOrder.DescendingOrder)
        else:
            hdr.setSortIndicator(-1, Qt.SortOrder.AscendingOrder)
        self._table.resizeRowsToContents()
        doc_label = "請求書" if doc_type == "invoice" else "領収書"
        if show_all:
            self._status_label.setText(
                f"{len(pm_data)} 件表示　（{doc_label}発行済 {issued_count} 件）")
        else:
            self._status_label.setText(
                f"未発行 {len(pm_data)} 件　／　{doc_label}発行済 {issued_count} 件")

    # ── 行数量取得 / キャッシュ保存 ──────────────────────────────

    def _get_row_quantities(self, row: int) -> dict[int, int]:
        result = {}
        for col_offset, tmpl in enumerate(self._templates):
            spin = self._table.cellWidget(row, 5 + col_offset * 2 + 1)
            if isinstance(spin, _QtySpinBox):
                result[tmpl["id"]] = spin.value()
        return result

    def _get_row_prices(self, row: int) -> dict[int, int]:
        result = {}
        for col_offset, tmpl in enumerate(self._templates):
            spin = self._table.cellWidget(row, 5 + col_offset * 2)
            if isinstance(spin, _QtySpinBox):
                result[tmpl["id"]] = spin.value()
        return result

    def _save_qty_cache(self):
        for r in range(self._table.rowCount()):
            data_item = self._table.item(r, COL_ORG)
            if not data_item:
                continue
            pm_id, _, _ = data_item.data(Qt.ItemDataRole.UserRole)
            self._qty_cache[pm_id] = self._get_row_quantities(r)
            self._price_cache[pm_id] = self._get_row_prices(r)

    # ── Excel入出力 ───────────────────────────────────────────────

    _XLSX_FIXED_HEADERS = ["ID", "会員番号", "事業所名", "フリガナ", "代表者名"]

    def _export_excel(self):
        """表示中の名簿＋項目ごとの単価・数量をExcelに出力する。"""
        project_id = self._proj_combo.currentData()
        if project_id is None:
            QMessageBox.information(self, "案件未選択", "件名を選択してください。")
            return
        if not self._templates:
            QMessageBox.information(
                self, "項目なし",
                "この案件には項目テンプレートが設定されていません。")
            return
        if self._table.rowCount() == 0:
            QMessageBox.information(self, "対象なし", "表示中の名簿がありません。")
            return

        proj_name = self._proj_combo.currentText()
        safe = "".join(c for c in proj_name if c not in '\\/:*?"<>|')
        path, _ = QFileDialog.getSaveFileName(
            self, "Excel出力", f"{safe}_単価数量入力.xlsx", "Excel (*.xlsx)")
        if not path:
            return

        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "単価数量入力"

        # ヘッダー：項目ごとに「単価」「数量」の2列
        headers = self._XLSX_FIXED_HEADERS + [
            col
            for t in self._templates
            for col in [f"{t['name']}（単価）", f"{t['name']}（数量）"]
        ]
        ws.append(headers)

        fill_fixed = PatternFill("solid", fgColor="DDEBF7")   # 固定列：青
        fill_price = PatternFill("solid", fgColor="FCE4D6")   # 単価列：橙
        fill_qty   = PatternFill("solid", fgColor="E2EFDA")   # 数量列：緑
        n_fixed = len(self._XLSX_FIXED_HEADERS)
        for ci, cell in enumerate(ws[1]):
            cell.font = Font(bold=True)
            pos = ci - n_fixed
            if ci < n_fixed:
                cell.fill = fill_fixed
            elif pos % 2 == 0:
                cell.fill = fill_price
            else:
                cell.fill = fill_qty

        exported = 0
        for r in range(self._table.rowCount()):
            data_item = self._table.item(r, COL_ORG)
            if not data_item:
                continue
            pm_id, _, _ = data_item.data(Qt.ItemDataRole.UserRole)
            qty   = self._get_row_quantities(r)
            price = self._get_row_prices(r)
            row_vals = [
                pm_id,
                self._table.item(r, COL_NUM).text(),
                self._table.item(r, COL_ORG).text(),
                self._table.item(r, COL_KANA).text(),
                self._table.item(r, COL_REP).text(),
            ]
            for t in self._templates:
                row_vals.append(price.get(t["id"], t["unit_price"]))
                row_vals.append(qty.get(t["id"], 0))
            ws.append(row_vals)
            exported += 1

        widths = [6, 10, 28, 22, 14] + [10, 10] * len(self._templates)
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        try:
            wb.save(path)
        except PermissionError:
            QMessageBox.critical(
                self, "保存エラー",
                "ファイルを保存できませんでした。\n"
                "同じファイルをExcelで開いたままになっていないか確認してください。")
            return
        QMessageBox.information(
            self, "Excel出力",
            f"{exported}件を出力しました。\n{path}\n\n"
            "Excelで単価・数量を編集後、「Excel取込」で読み込んでください。\n"
            "・数量0の項目は明細に含まれません\n"
            "・全項目0の行は発行対象外になります\n"
            "・ID列は照合に使うため変更しないでください")

    def _import_excel(self):
        """Excel出力で編集したファイルを読み込み、数量とチェックを画面に反映する。"""
        if self._table.rowCount() == 0:
            QMessageBox.information(self, "対象なし", "表示中の名簿がありません。")
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "Excel取込", "", "Excel (*.xlsx)")
        if not path:
            return

        import openpyxl
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
            wb.close()
        except Exception as e:
            QMessageBox.critical(self, "読込エラー", str(e))
            return
        if not all_rows:
            QMessageBox.warning(self, "読込エラー", "データが見つかりませんでした。")
            return

        header = [str(c).strip() if c is not None else "" for c in all_rows[0]]
        if "ID" not in header:
            QMessageBox.critical(
                self, "読込エラー",
                "見出し行に「ID」列が見つかりません。\n"
                "「Excel出力」で出力したファイルを使用してください。")
            return
        id_col = header.index("ID")

        # 数量列・単価列を検出（新フォーマット優先、旧フォーマットも対応）
        qty_cols: dict[int, int] = {}    # {tmpl_id: col_index}
        price_cols: dict[int, int] = {}  # {tmpl_id: col_index}
        missing_cols: list[str] = []
        for t in self._templates:
            new_qty_key   = f"{t['name']}（数量）"
            new_price_key = f"{t['name']}（単価）"
            if new_qty_key in header:
                qty_cols[t["id"]]   = header.index(new_qty_key)
                if new_price_key in header:
                    price_cols[t["id"]] = header.index(new_price_key)
            elif t["name"] in header:
                # 旧フォーマット（テンプレート名のみ = 数量列）
                qty_cols[t["id"]] = header.index(t["name"])
            else:
                missing_cols.append(t["name"])

        if not qty_cols:
            QMessageBox.critical(
                self, "読込エラー",
                "この案件の項目に対応する数量列が1つも見つかりません。\n"
                "案件の選択が出力時と同じか確認してください。")
            return

        def _parse_int(v, label: str, lo: int, hi: int,
                       bad: list[str], row_no: int) -> int | None:
            if v is None or str(v).strip() == "":
                return None
            try:
                f = float(str(v).strip())
                iv = int(f)
                if iv < lo:
                    raise ValueError
                if iv != f:
                    bad.append(f"{row_no}行目: {label}「{v}」は整数でないため{iv}として扱います")
                if iv > hi:
                    bad.append(f"{row_no}行目: {label}「{v}」は上限の{hi}に丸めました")
                    iv = hi
                return iv
            except (ValueError, OverflowError):
                bad.append(f"{row_no}行目: {label}「{v}」が不正です（スキップ）")
                return None

        file_qty:   dict[int, dict[int, int]] = {}
        file_price: dict[int, dict[int, int]] = {}
        bad_rows: list[str] = []
        for i, cells in enumerate(all_rows[1:], start=2):
            raw_id = cells[id_col] if id_col < len(cells) else None
            if raw_id is None or str(raw_id).strip() == "":
                continue
            try:
                pm_id = int(str(raw_id).strip())
            except ValueError:
                bad_rows.append(f"{i}行目: ID「{raw_id}」が数値ではありません")
                continue
            q = {}
            for tid, col in qty_cols.items():
                v = cells[col] if col < len(cells) else None
                iv = _parse_int(v, "数量", 0, 9999, bad_rows, i)
                q[tid] = iv if iv is not None else 0
            file_qty[pm_id] = q

            p = {}
            for tid, col in price_cols.items():
                v = cells[col] if col < len(cells) else None
                iv = _parse_int(v, "単価", 0, 9_999_999, bad_rows, i)
                if iv is not None:
                    p[tid] = iv
            if p:
                file_price[pm_id] = p

        applied = 0
        checked = 0
        for r in range(self._table.rowCount()):
            data_item = self._table.item(r, COL_ORG)
            if not data_item:
                continue
            pm_id, _, _ = data_item.data(Qt.ItemDataRole.UserRole)
            if pm_id not in file_qty:
                continue
            q = file_qty.pop(pm_id)
            p = file_price.pop(pm_id, {})
            for col_offset, tmpl in enumerate(self._templates):
                price_spin = self._table.cellWidget(r, 5 + col_offset * 2)
                qty_spin   = self._table.cellWidget(r, 5 + col_offset * 2 + 1)
                if isinstance(price_spin, _QtySpinBox) and tmpl["id"] in p:
                    price_spin.setValue(p[tmpl["id"]])
                if isinstance(qty_spin, _QtySpinBox) and tmpl["id"] in q:
                    qty_spin.setValue(q[tmpl["id"]])
            total = sum(q.values())
            chk = self._table.item(r, COL_CHK)
            if chk:
                chk.setCheckState(Qt.CheckState.Checked if total > 0
                                  else Qt.CheckState.Unchecked)
            if total > 0:
                checked += 1
            applied += 1
        self._save_qty_cache()

        has_price = bool(price_cols)
        msg = [f"{applied}件の{'単価・' if has_price else ''}数量を反映し、{checked}件に発行チェックを入れました。",
               "内容を確認のうえ「選択行に発行」ボタンで発行してください。"]
        if file_qty:
            msg.append(f"※名簿に表示されていないID {len(file_qty)}件は反映できませんでした。\n"
                       "（発行済み等で非表示の可能性。表示を「すべて」にして再度取り込んでください）")
        if missing_cols:
            msg.append("※次の項目の数量列が見つかりませんでした：" + "、".join(missing_cols))
        if bad_rows:
            shown = "\n".join(bad_rows[:5])
            more = f"\n…ほか{len(bad_rows) - 5}件" if len(bad_rows) > 5 else ""
            msg.append(f"※読み込めなかった値：\n{shown}{more}")
        QMessageBox.information(self, "Excel取込", "\n\n".join(msg))

    # ── チェック済み行の取得 ──────────────────────────────────────

    def _checked_rows(self) -> list[tuple[int, tuple]]:
        result = []
        for r in range(self._table.rowCount()):
            chk = self._table.item(r, COL_CHK)
            if chk and chk.checkState() == Qt.CheckState.Checked:
                data_item = self._table.item(r, COL_ORG)
                if data_item:
                    result.append((r, data_item.data(Qt.ItemDataRole.UserRole)))
        return result

    # ── 発行処理 ──────────────────────────────────────────────────

    def _do_issue_rows(self, rows: list[tuple[int, tuple]]) -> list[str]:
        """rows = [(row_idx, (pm_id, inv_id, rcp_id)), ...] を発行して PDF 生成。
        エラーメッセージのリストを返す。

        支払期限ダイアログはDB変更前に表示し、キャンセル時は何も変更しない。
        PDF生成に失敗した行は発行済みを取り消して「準備中」に戻す。
        """
        project_id = self._proj_combo.currentData()
        doc_type = self._doc_type
        delivery = self._delivery_combo.currentText()
        errors = []

        # ── 対象行の事前確定（DB変更前）─────────────────────────
        targets = []
        for row_idx, (pm_id, invoice_id, receipt_id) in rows:
            if doc_type == "invoice" and invoice_id is None and receipt_id is not None:
                continue  # 無効化済み
            quantities = self._get_row_quantities(row_idx)
            unit_prices = self._get_row_prices(row_idx)
            issuance_id = invoice_id if doc_type == "invoice" else receipt_id
            if issuance_id is None and quantities and not any(quantities.values()):
                org_item = self._table.item(row_idx, COL_ORG)
                name = org_item.text() if org_item else f"{row_idx + 1}行目"
                errors.append(f"{name}：数量がすべて0のためスキップしました")
                continue
            targets.append((pm_id, issuance_id, quantities, unit_prices))
        if not targets:
            return errors

        # ── 支払期限・封筒オプション（請求書のみ）/ 発行日（領収書のみ）──
        due_date = None
        window_envelope = False
        show_recipient_person = True
        receipt_issued_at = None
        if doc_type == "invoice":
            qd = self._due_date.date()
            due_date = date(qd.year(), qd.month(), qd.day())
            window_envelope = self._window_envelope_chk.isChecked()
            show_recipient_person = self._show_person_chk.isChecked()
        else:
            from datetime import datetime as _dt
            qd = self._issued_date.date()
            receipt_issued_at = _dt(qd.year(), qd.month(), qd.day())

        # ── 保存先フォルダを選択（メール送付以外）──────────────────
        save_dir: str | None = None
        if delivery != "メール送付":
            from app.utils.pdf_helpers import get_pdf_output_dir
            save_dir = QFileDialog.getExistingDirectory(
                self, "PDFの保存先フォルダを選択", get_pdf_output_dir()
            )
            if not save_dir:
                return errors  # キャンセル → DB変更なしで終了

        # ── 発行 → PDF生成（失敗時は発行を取り消す）──────────────
        session = get_session()
        issued_issuances = []
        pdf_paths = []
        open_each = len(targets) == 1 and delivery != "メール送付"
        try:
            from app.database.models import ProjectMember, Issuance
            from app.utils.pdf_helpers import generate_and_open, merge_and_open
            for pm_id, issuance_id, quantities, unit_prices in targets:
                try:
                    pm = session.get(ProjectMember, pm_id)
                    if issuance_id is None:
                        today = date.today()
                        iss = create_issuance_for_member(
                            session, project_id=project_id,
                            project_member_id=pm_id,
                            recipient_organization=pm.organization_name,
                            recipient_name=pm.representative_name,
                            recipient_department=pm.department or "",
                            doc_type=doc_type,
                            fiscal_year=today.year, month=today.month,
                            quantities=quantities if quantities else None,
                            unit_prices=unit_prices if unit_prices else None,
                            show_recipient_person=show_recipient_person,
                        )
                        issuance_id = iss.id

                    iss = session.get(Issuance, issuance_id)
                    if iss is None:
                        continue
                    was_issued = iss.status == "発行済み"
                    if not was_issued:
                        mark_as_issued(session, issuance_id,
                                       staff_id=current_user.get_id(),
                                       staff_name=current_user.get_name(),
                                       delivery_method=delivery,
                                       issued_at=receipt_issued_at)
                        iss = session.get(Issuance, issuance_id)
                        from app.services.operation_log_service import add_log
                        _lbl = "請求書" if doc_type == "invoice" else "領収書"
                        add_log(session, "発行", "issuance", issuance_id,
                                f"{_lbl} {iss.doc_number} 宛先：{iss.recipient_organization or iss.recipient_name}")
                    elif (iss.delivery_method or "") != delivery:
                        # 発行済み行の再実行時も配付方法を実態に合わせる
                        iss.delivery_method = delivery
                        session.commit()

                    try:
                        from app.database.models import Project as _Project
                        _proj = session.get(_Project, iss.project_id)
                        _pdf_save_path = (
                            os.path.join(save_dir, f"{iss.doc_number}.pdf")
                            if save_dir else None
                        )
                        path = generate_and_open(iss, session, due_date=due_date,
                                                 open_file=open_each,
                                                 save_path=_pdf_save_path,
                                                 window_envelope=window_envelope,
                                                 project=_proj)
                        if path:
                            pdf_paths.append(path)
                    except Exception as e:
                        if not was_issued:
                            # PDFが作れなかった行は発行前の状態に戻す
                            iss.status = "準備中"
                            iss.issued_at = None
                            session.commit()
                        name = (iss.recipient_organization
                                or iss.recipient_name or iss.doc_number)
                        errors.append(
                            f"{name}：PDF生成に失敗したため発行を取り消しました（{e}）")
                        continue
                    issued_issuances.append((iss, session))
                except Exception as e:
                    errors.append(str(e))

            if delivery == "メール送付" and issued_issuances:
                self._send_issue_emails(issued_issuances, errors)
            elif len(pdf_paths) > 1:
                # 一括発行：個別に開かず1つに結合して開く（連続印刷用）
                try:
                    merge_and_open(pdf_paths, self._proj_combo.currentText(), output_dir=save_dir)
                except Exception as e:
                    errors.append(f"PDF結合に失敗しました：{e}")
        finally:
            session.close()
        return errors

    def _send_issue_emails(self, issued_issuances: list, errors: list[str]):
        """配付方法「メール送付」で発行した分のPDFを1件ずつ確認してM365で送信する。"""
        from PyQt6.QtCore import QThread
        from PyQt6.QtWidgets import QApplication, QProgressDialog, QDialog
        from app.database.models import ProjectMember
        from app.services.email_service import prepare_issuance_email
        from app.services.operation_log_service import add_log
        from app.ui.invoice_mail_confirm_dialog import InvoiceMailConfirmDialog
        from app.ui.m365_mail_worker import M365MailWorker
        from app.utils.app_config import get_m365_client_id, get_m365_tenant_id
        label = "請求書" if self._doc_type == "invoice" else "領収書"

        client_id = get_m365_client_id()
        tenant_id = get_m365_tenant_id()
        if not client_id or not tenant_id:
            QMessageBox.critical(
                self, "設定エラー",
                "Microsoft 365 の Client ID / Tenant ID が設定されていません。\n"
                "設定 → メール設定から入力してください。")
            return

        sent = 0
        for iss, sess in issued_issuances:
            email = ""
            if iss.project_member_id:
                pm = sess.get(ProjectMember, iss.project_member_id)
                email = (pm.email or "").strip() if pm else ""

            try:
                to_addr, subject, body_html, pdf_path = prepare_issuance_email(
                    sess, iss, to_addr=email or None)
            except Exception as prep_err:
                errors.append(f"メール準備失敗：{prep_err}")
                add_log(sess, "メール送信失敗", "issuance", iss.id,
                        f"{label} {iss.doc_number}：{prep_err}")
                continue

            dlg = InvoiceMailConfirmDialog(
                self,
                to_recipients=[to_addr],
                subject=subject,
                body_html=body_html,
                pdf_path=pdf_path,
                invoice_no=iss.doc_number,
                customer_name=(iss.recipient_organization
                               or iss.recipient_name or ""),
                amount_text=f"¥{iss.amount:,}" if iss.amount else "",
            )
            if dlg.exec() != QDialog.DialogCode.Accepted:
                continue

            thread = QThread(self)
            worker = M365MailWorker(
                client_id, tenant_id, [to_addr],
                subject, body_html, pdf_path)
            worker.moveToThread(thread)
            prog = QProgressDialog(
                f"送信中（{iss.doc_number}）…", None, 0, 0, self)
            prog.setWindowTitle("メール送信")
            prog.setWindowModality(Qt.WindowModality.WindowModal)
            prog.show()
            _result: dict = {}
            def _on_done(r, _r=_result, _t=thread):
                _r["ok"] = r
                _t.quit()
            def _on_err(msg, _r=_result, _t=thread):
                _r["err"] = msg
                _t.quit()
            worker.finished.connect(_on_done)
            worker.failed.connect(_on_err)
            thread.started.connect(worker.run)
            thread.finished.connect(prog.close)
            thread.finished.connect(thread.deleteLater)
            thread.start()
            while thread.isRunning():
                QApplication.processEvents()

            if "ok" in _result:
                sent += 1
                add_log(sess, "メール送信", "issuance", iss.id,
                        f"{label} {iss.doc_number} → {to_addr}")
            else:
                err_msg = _result.get("err", "不明なエラー")
                errors.append(f"メール送信失敗：{err_msg}")
                add_log(sess, "メール送信失敗", "issuance", iss.id,
                        f"{label} {iss.doc_number}：{err_msg}")

        if sent > 0:
            QMessageBox.information(
                self, "メール送信", f"{sent}件のメールを送信しました。")

    def _issue_checked(self):
        targets = self._checked_rows()
        if not targets:
            QMessageBox.information(self, "未選択",
                                    "発行する行のチェックボックスにチェックを入れてください。")
            return
        errors = self._do_issue_rows(targets)
        if errors:
            QMessageBox.critical(self, "PDF生成エラー", "\n".join(errors))
        self._load_members()

    def _issue_all(self):
        project_id = self._proj_combo.currentData()
        if project_id is None:
            return
        label = "請求書" if self._doc_type == "invoice" else "領収書"
        ans = QMessageBox.question(
            self, "確認",
            f"表示中の全員ぶんを{label}で発行します。よろしいですか？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if ans != QMessageBox.StandardButton.Yes:
            return
        all_rows = []
        for r in range(self._table.rowCount()):
            data_item = self._table.item(r, COL_ORG)
            if data_item:
                all_rows.append((r, data_item.data(Qt.ItemDataRole.UserRole)))
        errors = self._do_issue_rows(all_rows)
        if errors:
            QMessageBox.critical(self, "PDF生成エラー", "\n".join(errors))
        self._load_members()
